import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import os

def process_shift_data(input_path, output_path):
    """
    Complete runnable script to process shift schedule Excel data.
    Creates a fresh Excel file with the same data.
    
    Args:
        input_path (str): Path to input Excel file
        output_path (str): Path for output Excel file
    
    Returns:
        bool: True if successful, False otherwise
    """
    try:
        # Verify input file exists
        if not os.path.exists(input_path):
            raise FileNotFoundError(f"Input file not found: {input_path}")
        
        # Read Excel file (automatically handles multiple sheets)
        excel_data = pd.ExcelFile(input_path)
        
        # Create fresh workbook
        wb = Workbook()
        
        # Process each sheet
        for sheet_name in excel_data.sheet_names:
            df = pd.read_excel(excel_data, sheet_name=sheet_name)
            
            # Create new sheet (remove default Sheet if it exists)
            if sheet_name == "Sheet" and "Sheet" in wb.sheetnames:
                ws = wb["Sheet"]
                ws.title = sheet_name
            else:
                ws = wb.create_sheet(title=sheet_name)
            
            # Write data to sheet
            for r in dataframe_to_rows(df, index=False, header=True):
                ws.append(r)
        
        # Remove default empty sheet if it exists
        if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
            wb.remove(wb["Sheet"])
        
        # Save the workbook
        wb.save(output_path)
        print(f"Success! New file created at: {output_path}")
        return True
        
    except Exception as e:
        print(f"Error: {str(e)}")
        return False

if __name__ == "__main__":
    # Example usage - replace with your actual file paths
    input_file = "shift_schedule.xlsx"  # Your input file
    output_file = "fresh_shift_data.xlsx"  # Output file
    
    # Run the processor
    success = process_shift_data(input_file, output_file)
    
    if success:
        print("Processing completed successfully!")
    else:
        print("Processing failed. Please check the error message.")